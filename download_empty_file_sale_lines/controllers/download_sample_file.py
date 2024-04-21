from odoo import http
from odoo.http import request
from odoo.addons.web.controllers.main import content_disposition
import base64
import xlsxwriter
from openpyxl import load_workbook
from odoo.modules.module import get_resource_path

class Download_xls(http.Controller):
    
    @http.route('/web/binary/download_document', type='http', auth="public")
    def download_document(self,model,id, **kw):
         

        sale_order_line_xls = request.env['ir.attachment'].search([('name','=','MetraQuotationTemplate.xlsx')])
        filecontent = sale_order_line_xls.datas
        filename = 'MetraQuotationTemplate.xlsx'
        filecontent = base64.b64decode(filecontent)

        return request.make_response(filecontent,
            [('Content-Type', 'application/octet-stream'),
            ('Content-Disposition', content_disposition(filename))])
         
      

   



      



    # @http.route('/web/binary/download_document_with_data', type='http', auth="public")
    # def download_document_with_data(self,model,id, **kw):



    #     excel = get_resource_path('download_empty_file_sale_lines', 'sample_file', 'MetraQuotationTemplate.xlsx')
    #     print('------------------------------------------ppppppppppppppaaaaaaaaaattttttttttttttt', excel)
    #     workbook = load_workbook(filename=excel)        
    #     sheet = workbook['Import Sale Order Lines']
    #     new_workbook = xlsxwriter.Workbook('MetraQuotationTemplate')
    #     new_sheet= new_workbook.add_worksheet('Sale Order Lines')
    #     format_1 = new_workbook.add_format({'bold': True, 'align': 'center', 'valign': 'vcenter', 'bg_color': '#cccccc'})
    #     bold = new_workbook.add_format({'bold': True,'valign': 'vcenter'})


    #     new_sheet.set_row(0, 30, format_1)
    #     new_sheet.set_column(0, 0, 30)
    #     new_sheet.set_column(1, 1, 30)
    #     new_sheet.set_column(2, 2, 30)
    #     new_sheet.set_column(3, 3, 30)
    #     new_sheet.set_column(4, 4, 30)
    #     new_sheet.set_column(5, 5, 30)
    #     new_sheet.set_column(6, 6, 30)
    #     new_sheet.set_column(7, 7, 30)
    #     new_sheet.set_column(8, 8, 30)
    #     new_sheet.set_column(9, 9, 30)
    #     new_sheet.set_column(10, 10, 30)
    #     new_sheet.set_column(11, 11, 30)
    #     new_sheet.set_column(12, 12, 30)
    #     new_sheet.set_column(13, 13, 30)
    #     new_sheet.set_column(14, 14, 30)
    #     new_sheet.set_column(15, 15, 30)
    #     new_sheet.set_column(16, 16, 30)
    #     new_sheet.set_column(17, 17, 30)
    #     new_sheet.set_column(18, 18, 30)
    #     new_sheet.set_column(19, 19, 30)
    #     new_sheet.set_column(20, 20, 30)

    #     order_id = request.env['sale.order'].search([('id','=',kw['order_id'])])


    #     duration_values = []

    #     for line in order_id.order_line:
    #         if isinstance(line.duration, str):
    #             try:
    #                 duration = int(line.duration)
    #             except ValueError:
    #                 line.duration = False

    #             else:
    #                 if duration != 0:
    #                     duration_values.append(duration)
    #                 elif duration == 0:
    #                     line.duration = False    


     


    #     for row_index, row in enumerate(sheet.iter_rows(values_only=True,), start=0):
    #         for col_index, value in enumerate(row, start=0):

    #             if isinstance(value, str) and value.startswith('='):
    #                 new_sheet.write_formula(row_index, col_index, f'{value}')
    #             else:
    #                 new_sheet.write(row_index, col_index, value)
          


    #     for index, line in enumerate(order_id.order_line):
               
    #                 if line.display_type == 'line_section':
    #                     new_sheet.merge_range(index+1,0,index+1,19, line.name,bold)
    #                 elif line.display_type == 'line_note':
    #                     new_sheet.merge_range(index+1,0,index+1,19, line.name,bold)    
    #                 else: 
                      
    #                         new_sheet.write(index+1, 0, line.line_number or index+1)  
    #                         new_sheet.write(index+1, 1, line.product_template_id.name)
    #                         new_sheet.write(index+1, 2, line.smart_account_mandatory or "--")
    #                         new_sheet.write(index+1, 3, line.name)
    #                         new_sheet.write(index+1, 4, line.cisco_product_ref or "--")
    #                         new_sheet.write(index+1, 5, line.product_family or "--")
    #                         new_sheet.write(index+1, 6, line.duration or 'N/A')
    #                         new_sheet.write(index+1, 7, line.estimated_lead_time or 'N/A')
    #                         new_sheet.write(index+1, 8, line.cost)
    #                         new_sheet.write(index+1, 9, line.pricing_term or "--")
    #                         new_sheet.write(index+1, 10, line.product_uom_qty)
    #                         row_num = index+2
    #                         new_sheet.write_formula(index+1, 11, f'=ROUND(I{row_num}-(I{row_num}*(M{row_num}/100)),2)')
    #                         new_sheet.write(index+1, 12, line.discount_metra)
    #                         new_sheet.write_formula(index+1, 13, f'=ROUND((K{row_num} * L{row_num}),2)')
    #                         new_sheet.write_formula(index+1, 14, f'=ROUND(I{row_num}-(I{row_num}*(P{row_num}/100)),2)')
    #                         new_sheet.write(index+1, 15, line.partner_discount)
    #                         new_sheet.write(index+1, 16, line.conditions)
    #                         new_sheet.write(index+1, 17, line.mergin)
    #                         new_sheet.write_formula(index+1, 18, f'=(L{row_num}*(1+Q{row_num}/100))/(1-R{row_num}/100)')
    #                         new_sheet.write_formula(index+1, 19, f'=S{row_num}*K{row_num}')
                 
        

    #     new_workbook.close() 


    #     with open('MetraQuotationTemplate', 'rb') as file:
    #         filecontent = file.read()



    #     filename = 'MetraQuotationTemplate.xlsx'


    #     return request.make_response(filecontent,
    #         [('Content-Type', 'application/octet-stream'),
    #         ('Content-Disposition', content_disposition(filename))])